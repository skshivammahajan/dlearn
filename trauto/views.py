from django.core.mail import send_mail
from django.core.paginator import Paginator , PageNotAnInteger , EmptyPage
from django.http import HttpResponse
from django.shortcuts import render
import openpyxl
from django.contrib.auth.decorators import login_required

from ericauto import settings
from trauto.models import NodeInventory


@login_required(login_url='/user/login/')
def index(request):
    if "GET" == request.method:
        return render(request, 'file_upload.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        # sheets = wb.sheetnames
        # getting active sheet
        # active_sheet = wb.active
        # reading a cell
        # print ( worksheet [ "A1" ].value )

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        print(excel_data)

        for data in excel_data:
            node = NodeInventory(node_name=data[-1], node_ip=data[1], oss_ip=data[0])
            node.save()
        return render(request, 'file_upload.html', {"excel_data": excel_data})


def mail(request):
    subject = "Greetings"
    msg = "Congratulations for your success"
    to = "shivam.b.gupta@ericsson.com"
    res = send_mail(subject, msg, settings.EMAIL_HOST_USER, [to])
    if res == 1:
        msg = "Mail Sent Successfuly"
    else:
        msg = "Mail could not sent"
    return HttpResponse(msg)


def node_list(request):
    user_list = NodeInventory.objects.all()
    page = request.GET.get('page', 1)

    paginator = Paginator(user_list, 20)

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)
    print ( users.paginator.page_range )
    return render(request, 'node_list.html', { 'users': users })